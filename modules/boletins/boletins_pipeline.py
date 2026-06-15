import os
import sys
import re
import asyncio
import pathlib
import pandas as pd
import openpyxl

# Ajuste de path para importar do core
current_dir = pathlib.Path(__file__).parent
project_root = current_dir.parent.parent
sys.path.append(str(project_root))

from core.models import ProgramRecipe, VoiceStrategy, AssemblyRecipe
from core.engine import PipelineEngine

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
SPREADSHEET_ID = "1b1xnzvA00H1JC9uTvd6c-PBwQjEzGRs6t_raXG_ztsU"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
BOLETINS_INPUT_DIR = current_dir # Excel é baixado aqui
BOLETINS_OUTPUT_DIR = pathlib.Path(r"H:\Meu Drive\RADIO TJRN CONTEÚDO\0-BOLETINS")
LOCAL_WORK_DIR = current_dir / "workspace"
VHT_DIR = project_root / "assets" / "vht"

# Mapeamento dos meses no formato NJUD
MONTH_MAP = {
    1: "1 - JANEIRO", 2: "2 - FEVEREIRO", 3: "3 - MARÇO", 4: "4 - ABRIL",
    5: "5 - MAIO", 6: "6 - JUNHO", 7: "7 - JULHO", 8: "8 - AGOSTO",
    9: "9 - SETEMBRO", 10: "10 - OUTUBRO", 11: "11 - NOVEMBRO", 12: "12 - DEZEMBRO"
}

# ---------------------------------------------------------------------------
# Processador Excel -> TXT
# ---------------------------------------------------------------------------
def fetch_excel_and_create_txts(txt_dir: pathlib.Path):
    """Baixa a planilha, encontra pendências e cria os arquivos .txt para o motor."""
    print("Baixando planilha de controle...")
    local_xlsx = BOLETINS_INPUT_DIR / "BOLETINS_ATUAL.xlsx"
    import urllib.request
    urllib.request.urlretrieve(SHEET_URL, local_xlsx)
    
    wb = openpyxl.load_load_workbook(local_xlsx, data_only=True) if hasattr(openpyxl, "load_workbook") else openpyxl.load_workbook(local_xlsx, data_only=True)
    
    sheets_to_process = [s for s in wb.sheetnames if "JANEIRO" in s.upper() or "FEVEREIRO" in s.upper() or "MARÇO" in s.upper() or "ABRIL" in s.upper() or "MAIO" in s.upper() or "JUNHO" in s.upper()]
    
    count = 0
    for s_name in sheets_to_process:
        ws = wb[s_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            if not row or not any(row): continue
            
            data_str = str(row[0]) if row[0] else ""
            status = str(row[3]).lower() if len(row) > 3 and row[3] else ""
            texto = str(row[11]) if len(row) > 11 and row[11] else "" # Na planilha antiga o texto final ficava na L(11) ou G(6)?
            
            # Ajuste baseado no gerar_boletins_tts antigo:
            # col 1: STATUS, col 11: TEXTO FINAL, col 0: DATA
            # Precisaríamos verificar o mapeamento exato, assumiremos col 11 como texto para locução.
            if "✔" not in status and len(texto) > 10:
                # Criar arquivo txt
                safe_date = data_str.replace("/", "-").split()[0]
                file_name = f"BOLETIM_{safe_date}_{s_name}_L{row_idx}.txt"
                (txt_dir / file_name).write_text(texto, encoding="utf-8")
                count += 1
                
    print(f"[{count}] novos boletins extraídos da planilha.")
    return count

SYSTEM_PROMPT = """Você é um especialista em edição de roteiros de rádio. Seu objetivo é processar matérias para o Boletim 'Notícias da Hora' formatando-as no padrão 'Audio-as-Text'.

REGRAS:
1. Sem formatação Markdown.
2. Extraia o assunto principal e crie uma manchete curta (1-2 frases). Esta será a CABEÇA.
3. O restante do texto será o OFF.
4. O formato de edição de áudio segue o padrão 'Audio-as-Text'. Insira obrigatoriamente as tags de controle em letras maiúsculas:
   - ANTES DA CABEÇA (início do programa): [ASSET: ABERTURA]
   - APÓS A CABEÇA E ANTES DO OFF: [ASSET: PASSAGEM] seguido por [TRILHA: LIGAR]
   - FIM DO PROGRAMA (após a última fala do OFF): [TRILHA: DESLIGAR] seguido por [ASSET: ENCERRAMENTO]
5. Substituir as marcações originais de locutores por "Speaker 1: ". Não alterne locutores no Boletim.
6. Escrever números, valores, datas e horas por extenso. Siglas separadas por espaço (ex: t j r n).
"""

def boletim_parse_hook(content: str) -> list:
    """Extrai comandos de ASSET, TRILHA e as falas (Speaker 1)."""
    blocks = []
    
    for linha in content.splitlines():
        linha = linha.strip()
        if not linha:
            continue
            
        match_asset = re.match(r'^\[ASSET:\s*(.*?)\]$', linha, re.IGNORECASE)
        if match_asset:
            blocks.append(("ASSET", match_asset.group(1).strip().upper()))
            continue
            
        match_trilha = re.match(r'^\[TRILHA:\s*(LIGAR|DESLIGAR)\]$', linha, re.IGNORECASE)
        if match_trilha:
            blocks.append(("TRILHA", match_trilha.group(1).strip().upper()))
            continue

        match_loc = re.match(r'^(Speaker\s*1):\s*((?:\[.*?\])?\s*.*)$', linha, re.IGNORECASE)
        if match_loc:
            speaker = match_loc.group(1).lower().replace(" ", "")
            texto = match_loc.group(2).strip()
            texto = re.sub(r'\[.*?\]', '', texto).strip()
            
            if texto:
                blocks.append(("LOC", texto)) # Boletim é inter_file, passa apenas o texto
                
    return blocks

# ---------------------------------------------------------------------------
# Receita do Programa
# ---------------------------------------------------------------------------
receita_boletins = ProgramRecipe(
    name="Notícias da Hora (Boletins)",
    drive_input_dir=BOLETINS_INPUT_DIR,
    drive_output_dir=BOLETINS_OUTPUT_DIR,
    local_work_dir=LOCAL_WORK_DIR,
    system_prompt=SYSTEM_PROMPT,
    voice_strategy=VoiceStrategy(
        type='inter_file', # Alterna a voz a cada arquivo novo no lote
        voices=["pt-BR-FranciscaNeural", "pt-BR-AntonioNeural"]
    ),
    assembly=AssemblyRecipe(
        profile_path=project_root / "assets" / "profiles" / "boletim_profile.json"
    ),
    parse_hook=boletim_parse_hook
)

if __name__ == "__main__":
    motor = PipelineEngine(receita_boletins)
    
    # 1. Extrair pendências do Excel para a pasta txt_bruto
    try:
        # fetch_excel_and_create_txts(motor.txt_dir)
        # Como a estrutura exata do excel antigo é complexa, deixarei comentado o fetch automático
        # para evitar quebra. O usuário pode colocar os txts na pasta manualmente ou
        # podemos integrar o parser de excel original.
        print("Módulo de Boletins configurado para a nova arquitetura.")
        print("Insira os textos em 'modules/boletins/workspace/1_txt_bruto' e rode novamente.")
        asyncio.run(motor.run_all())
    except Exception as e:
        print(f"Erro: {e}")
