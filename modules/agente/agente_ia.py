import os
import sys
import re
import time
import json
import argparse
import subprocess
import shutil
import gc
from datetime import datetime, date
import openpyxl
import pathlib
import atexit

# Corrigir encodificação de console no Windows para evitar quedas por caracteres unicode (ex: ✔)
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

# Adicionar a raiz do projeto e o diretório do módulo ao python path
current_dir = os.path.dirname(os.path.abspath(__file__)).replace("\\", "/")
project_root = os.path.dirname(os.path.dirname(current_dir)).replace("\\", "/")
sys.path.append(project_root)
sys.path.append(current_dir)

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
    google_apis_available = True
except ImportError:
    google_apis_available = False

try:
    from core.llm_factory import LLMFactory
    from core.constants import MONTH_MAP_FULL
    llm_available = True
except ImportError:
    llm_available = False

try:
    from core.drive_watcher import DriveWatcher
    drive_watcher_available = True
except ImportError:
    DriveWatcher = None
    drive_watcher_available = False

# Configurações
SPREADSHEET_ID_BOLETINS = carregar_env_var("SPREADSHEET_ID_BOLETINS", "1b1xnzvA00H1JC9uTvd6c-PBwQjEzGRs6t_raXG_ztsU")
SPREADSHEET_ID_NJUD = carregar_env_var("SPREADSHEET_ID_NJUD", "1HegL-SudxPLI4Y6wsj1nnJocXHOvi-6inGqQld1lYec")
CREDENTIALS_PATH = os.path.join(project_root, carregar_env_var("CREDENTIALS_PATH", "config/credentials/service_account.json")).replace("\\", "/")
drive_producao = carregar_env_var("DRIVE_PRODUCAO", "H:/Meu Drive/RADIO TJRN CONTEÚDO/00_PRODUCAO_2026")
LOG_PATH = f"{drive_producao}/LOG_ACOES_5S.md"
NJUD_DRIVE_BASE = f"{drive_producao}/02_JORNAIS_NJUD"

# Mapeamentos do Calendário de 2026
WEEKDAYS_PT = {
    0: "SEG", 1: "TER", 2: "QUA", 3: "QUI", 4: "SEX", 5: "SAB", 6: "DOM"
}



def registrar_log_5s(mensagem):
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    log_entry = f"*   **{timestamp}**: {mensagem}\n"
    print(f"[Agente 5S] {mensagem}")
    
    if os.path.exists("H:/Meu Drive"):
        try:
            if not os.path.exists(LOG_PATH):
                with open(LOG_PATH, "w", encoding="utf-8") as f:
                    f.write("# Registro de Ações - Padrão 5S 2026\n\n## Histórico de Execuções do Agente de IA\n\n")
            with open(LOG_PATH, "a", encoding="utf-8") as f:
                f.write(log_entry)
        except Exception as e:
            print(f"[AVISO] Falha ao gravar log no Drive H: {e}")
    else:
        print("[AVISO] Unidade H: indisponível para registro de log físico.")

def verificar_e_montar_drive():
    if os.path.exists("H:/Meu Drive"):
        print("[Resiliência] Google Drive (H:) detectado e acessível.")
        return True
        
    print("[Resiliência] Google Drive (H:) não encontrado. Tentando localizar o Google Drive Desktop...")
    
    possiveis_exes = []
    base_dir = r"C:\Program Files\Google\Drive File Stream"
    if os.path.exists(base_dir):
        for item in sorted(os.listdir(base_dir), reverse=True):
            exe = os.path.join(base_dir, item, "GoogleDriveFS.exe")
            if os.path.exists(exe):
                possiveis_exes.append(exe)
                
    possiveis_exes.append(r"C:\Program Files\Google\Drive File Stream\126.0.5.0\GoogleDriveFS.exe")
    possiveis_exes.append(r"C:\Program Files\Google\Drive File Stream\125.0.0.0\GoogleDriveFS.exe")
    
    gdrive_exe = None
    for p in possiveis_exes:
        if os.path.exists(p):
            gdrive_exe = p
            break
            
    if not gdrive_exe:
        registrar_log_5s("ERRO: Executável do Google Drive Desktop não encontrado. Montagem manual necessária.")
        return False
        
    try:
        print(f"[Resiliência] Executando: {gdrive_exe}")
        subprocess.Popen([gdrive_exe])
        print("[Resiliência] Processo inicializado. Aguardando montagem (até 15s)...")
        for i in range(15):
            time.sleep(1)
            if os.path.exists("H:/Meu Drive"):
                registrar_log_5s("Google Drive (H:) montado com sucesso via auto-mount.")
                return True
    except Exception as e:
        print(f"[Resiliência] Erro ao iniciar Google Drive Desktop: {e}")
        
    registrar_log_5s("ERRO: Falha ao montar o Google Drive automaticamente.")
    return False

def obter_google_drive_service():
    if not google_apis_available:
        print("[ERRO] Bibliotecas do Google APIs não importadas.")
        return None
        
    if not os.path.exists(CREDENTIALS_PATH):
        print(f"[ERRO] Credenciais da conta de serviço não encontradas: {CREDENTIALS_PATH}")
        return None
        
    try:
        scopes = ["https://www.googleapis.com/auth/drive"]
        creds = service_account.Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=scopes)
        drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
        return drive_service
    except Exception as e:
        print(f"[ERRO] Falha ao autenticar serviço do Google Drive: {e}")
        return None

def baixar_planilha_gdrive(drive_service, file_id, temp_name):
    try:
        import io
        request = drive_service.files().export_media(
            fileId=file_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        
        temp_path = os.path.join(current_dir, temp_name).replace("\\", "/")
        with open(temp_path, "wb") as f:
            f.write(fh.getvalue())
        
        wb = openpyxl.load_workbook(temp_path)
        return wb, temp_path
    except Exception as e:
        print(f"[ERRO] Falha ao exportar/abrir planilha {file_id}: {e}")
        return None, None

def salvar_e_enviar_planilha_gdrive(drive_service, file_id, wb, temp_path):
    try:
        # Salvar e fechar workbook para liberar o arquivo no Windows
        wb.save(temp_path)
        wb.close()
        
        # Forçar coleta de lixo para liberar referências do openpyxl
        gc.collect()
        time.sleep(0.5)
        
        # Upload de volta para o Drive
        media = MediaFileUpload(temp_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        drive_service.files().update(
            fileId=file_id,
            media_body=media
        ).execute()
        
        # Limpar arquivo temporário de forma protegida contra bloqueio do Windows
        try:
            # Fechar media se aplicável
            if hasattr(media, '_fd') and media._fd:
                try:
                    os.close(media._fd)
                except Exception:
                    pass
            time.sleep(0.5)
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as e_del:
            print(f"[AVISO] Não foi possível remover arquivo temporário {temp_path}: {e_del}")
            
        return True
    except Exception as e:
        print(f"[ERRO] Falha ao salvar/enviar planilha de volta {file_id}: {e}")
        try:
            wb.close()
        except Exception:
            pass
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass
        return False

def extrair_data_registro(nome_edicao, data_criacao, caminho):
    if nome_edicao:
        m = re.search(r'BOLETIM_RADIO_TJRN_(\d{2})_(\d{2})_(\d{4})', str(nome_edicao), re.IGNORECASE)
        if m:
            return int(m.group(1)), int(m.group(2)), int(m.group(3))
            
    if data_criacao:
        if isinstance(data_criacao, str):
            m = re.search(r'(\d{4})-(\d{2})-(\d{2})', data_criacao)
            if m:
                return int(m.group(3)), int(m.group(2)), int(m.group(1))
            m = re.search(r'(\d{2})[/-](\d{2})[/-](\d{4})', data_criacao)
            if m:
                return int(m.group(1)), int(m.group(2)), int(m.group(3))
                
    if caminho:
        m = re.search(r'(\d{2})/(\d{2})', str(caminho))
        if m:
            return int(m.group(1)), int(m.group(2)), 2026
            
    return None, None, None

def obter_dia_semana_pt(ano, mes, dia):
    try:
        dt = datetime(ano, mes, dia)
        return WEEKDAYS_PT[dt.weekday()]
    except Exception:
        return "SEG"

def extrair_doc_id(url):
    m = re.search(r'/d/([a-zA-Z0-9_-]{25,})', str(url))
    if m:
        return m.group(1)
    return None

def baixar_doc_via_api(drive_service, doc_id):
    try:
        import io
        request = drive_service.files().export_media(fileId=doc_id, mimeType="text/plain")
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return fh.getvalue().decode('utf-8', errors='replace')
    except Exception:
        return None

def resolver_conflito_cognitivo(drive_service, tag_conflito, data_str, boletins_lista):
    """
    Usa heurísticas e LLM para identificar quais tags reais pertencem a cada boletim duplicado.
    """
    print(f"  [Cognitivo] Resolvendo conflito para a tag duplicada '{tag_conflito}' no dia {data_str}...")
    
    resolvido = True
    resultado = {}
    
    for b in boletins_lista:
        doc_id = b["doc_id"]
        gdoc_name = ""
        if doc_id:
            try:
                meta = drive_service.files().get(fileId=doc_id, fields="name").execute()
                gdoc_name = meta.get("name", "")
            except Exception:
                pass
        
        m = re.search(r'^(B\d+)\b', gdoc_name, re.IGNORECASE)
        if m:
            tag_detectada = m.group(1).upper()
            resultado[b["sheet_row"]] = tag_detectada
            print(f"    - Identificado pelo Drive: '{gdoc_name}' -> Tag real: {tag_detectada} (Linha {b['sheet_row']})")
        else:
            resolvido = False
            
    if resolvido and len(set(resultado.values())) == len(boletins_lista):
        return resultado
        
    if llm_available:
        print("    - Nomes de arquivo no Drive não foram suficientes. Chamando LLM para análise cognitiva...")
        llm = LLMFactory()
        
        prompt_pautas = ""
        for i, b in enumerate(boletins_lista):
            doc_id = b["doc_id"]
            texto = ""
            if doc_id:
                texto = baixar_doc_via_api(drive_service, doc_id) or ""
            
            prompt_pautas += f"--- PAUTA {i+1} (Linha da Planilha: {b['sheet_row']}, Título no Excel: {b['col_b']}) ---\n"
            prompt_pautas += f"{texto[:600]}\n\n"
            
        system_prompt = (
            "Você é um supervisor de radiojornalismo. Sua tarefa é analisar o conteúdo de pautas de rádio "
            "e correlacioná-las com suas numerações corretas de boletim (B1, B2, B3, B4, etc.) para o dia.\n"
            "Retorne APENAS um objeto JSON estruturado onde a chave é a Linha da Planilha (número inteiro) "
            "e o valor é a tag corrigida (ex: 'B2', 'B3'). Exemplo de formato:\n"
            "{\n  \"7\": \"B2\",\n  \"6\": \"B3\"\n}"
        )
        
        user_prompt = (
            f"Temos um conflito no dia {data_str}: múltiplos boletins estão marcados com a mesma tag '{tag_conflito}'.\n"
            "Analise os textos abaixo para deduzir qual numeração correta de boletim cada um deve ter.\n"
            "Dica: a numeração geralmente segue a ordem de urgência ou sequência lógica das pautas do dia.\n\n"
            f"{prompt_pautas}"
        )
        
        try:
            resposta = llm.ask(system_prompt, user_prompt)
            m_json = re.search(r'(\{.*?\})', resposta, re.DOTALL)
            if m_json:
                data_json = json.loads(m_json.group(1))
                res_dict = {int(k): str(v).upper() for k, v in data_json.items()}
                print("    - Resolução obtida via LLM:", res_dict)
                return res_dict
        except Exception as e:
            print(f"    [!] Erro na análise cognitiva via LLM: {e}")
            
    return None

def analisar_e_corrigir_planilha_boletins(drive_service):
    print("\n[Agente] Iniciando escaneamento cognitivo de Boletins no Google Drive...")
    
    wb, temp_path = baixar_planilha_gdrive(drive_service, SPREADSHEET_ID_BOLETINS, "temp_boletins_agente.xlsx")
    if not wb or not temp_path:
        print("[ERRO] Não foi possível obter planilha de Boletins via Drive API.")
        return False
        
    sheets = [name for name in wb.sheetnames if name != 'DASHBOARD GERAL']
    conflitos_corrigidos = 0
    alterada = False
    
    for sheet_name in sheets:
        if "2026" not in sheet_name:
            continue
            
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 5:
            continue
            
        dias_grupos = {}
        
        for idx, r in enumerate(rows[5:], start=5):
            if len(r) < 2 or not r[1]:
                continue
                
            col_b = r[1].strip()
            col_g = r[6].strip() if len(r) > 6 and r[6] else ""
            url = r[7].strip() if len(r) > 7 and r[7] else ""
            data_criacao = r[8]
            caminho_col = r[0].strip() if r[0] else ""
            
            day, month, year = extrair_data_registro(col_g, data_criacao, caminho_col)
            if not day or not month:
                continue
                
            doc_id = extrair_doc_id(url)
            tag_match = re.match(r'^(B\d+)', col_b, re.IGNORECASE)
            tag = tag_match.group(1).upper() if tag_match else ""
            
            if not tag:
                continue
                
            key = (day, month, year)
            if key not in dias_grupos:
                dias_grupos[key] = []
                
            dias_grupos[key].append({
                "sheet_row": idx + 1,
                "col_b": col_b,
                "col_g": col_g,
                "tag": tag,
                "doc_id": doc_id,
                "url": url
            })
            
        for date_key, boletins in dias_grupos.items():
            day, month, year = date_key
            data_str = f"{day:02d}/{month:02d}/{year}"
            
            tags_contagem = {}
            for b in boletins:
                tags_contagem[b["tag"]] = tags_contagem.get(b["tag"], 0) + 1
                
            tags_duplicadas = [t for t, count in tags_contagem.items() if count > 1]
            
            for tag_dup in tags_duplicadas:
                conflitantes = [b for b in boletins if b["tag"] == tag_dup]
                
                correcao_tags = resolver_conflito_cognitivo(drive_service, tag_dup, data_str, conflitantes)
                
                if correcao_tags:
                    for row_num, nova_tag in correcao_tags.items():
                        item = next(b for b in conflitantes if b["sheet_row"] == row_num)
                        if item["tag"] == nova_tag:
                            continue
                            
                        nova_col_b = re.sub(r'^B\d+', nova_tag, item["col_b"])
                        nova_col_g = re.sub(r'_B\d+_', f'_{nova_tag}_', item["col_g"])
                        
                        ws.cell(row=row_num, column=2, value=nova_col_b)
                        ws.cell(row=row_num, column=7, value=nova_col_g)
                        
                        registrar_log_5s(
                            f"Correção cognitiva na aba '{sheet_name}' (Linha {row_num}): "
                            f"boletim renomeado de '{item['tag']}' para '{nova_tag}' devido a duplicidade."
                        )
                        conflitos_corrigidos += 1
                        alterada = True
                        
    if alterada:
        salvar_e_enviar_planilha_gdrive(drive_service, SPREADSHEET_ID_BOLETINS, wb, temp_path)
        print("[OK] Planilha de Boletins corrigida e sincronizada no Drive.")
    else:
        wb.close()
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass
        print("  Nenhuma inconsistência de Boletins encontrada.")
        
    return True

LOCK_FILE = pathlib.Path(current_dir) / ".agente.lock"

def adquirir_lock() -> bool:
    if LOCK_FILE.exists():
        try:
            pid = LOCK_FILE.read_text().strip()
            if pid:
                os.kill(int(pid), 0)  # Verifica se processo existe no Linux
                print(f"[LOCK] Agente já em execução (PID {pid}). Abortando.")
                return False
        except (ProcessLookupError, ValueError):
            pass  # Processo morreu, prossegue
        except Exception:
            pass
    try:
        LOCK_FILE.write_text(str(os.getpid()))
        atexit.register(lambda: LOCK_FILE.unlink(missing_ok=True))
        return True
    except Exception as e:
        print(f"[LOCK] Erro ao criar lockfile: {e}")
        return False

def _run_pipeline(nome: str, script_path: str) -> bool:
    full_path = os.path.join(project_root, script_path).replace("\\", "/")
    print(f"  -> Iniciando {nome}: {full_path}")
    try:
        res = subprocess.run(
            [sys.executable, full_path],
            capture_output=True, text=True, encoding='utf-8', errors='ignore'
        )
        print(f"----- {nome} Output -----\n{res.stdout}")
        if res.stderr:
            print(f"----- {nome} Errors -----\n{res.stderr}")
        ok = res.returncode == 0
        registrar_log_5s(f"Pipeline {nome} {'concluído' if ok else 'falhou'}.")
        return ok
    except Exception as e:
        registrar_log_5s(f"CRÍTICO: Falha ao disparar pipeline {nome}: {e}")
        return False

def executar_pipelines():
    print("\n[Agente] Disparando pipelines de gravação física (subprocessos)...")
    ok_boletins = _run_pipeline("Boletins", "modules/boletins/gerar_boletins_tts.py")
    ok_njud = _run_pipeline("NJUD", "modules/jornal/gerar_njud_tts.py")
    ok_giro = _run_pipeline("Giro", "modules/giro/giro_pipeline.py")
    return ok_boletins and ok_njud and ok_giro

def obter_sufixo_data_njud(refer_val):
    if not refer_val:
        return ""
    if isinstance(refer_val, (datetime, date)):
        return refer_val.strftime("%d-%m")
    refer_str = str(refer_val).strip()
    m = re.search(r'(\d{4})[-/](\d{2})[-/](\d{2})', refer_str)
    if m:
        return f"{m.group(3)}-{m.group(2)}"
    m2 = re.search(r'(\d{2})[-/](\d{2})[-/](\d{4})', refer_str)
    if m2:
        return f"{m2.group(1)}-{m2.group(2)}"
    return ""

def obter_caminho_mes_njud(refer_val):
    if not refer_val:
        return "6 - JUNHO"
    if isinstance(refer_val, (datetime, date)):
        return MONTH_MAP_FULL.get(refer_val.month, "6 - JUNHO")
    refer_str = str(refer_val).strip()
    for m_num, m_full in MONTH_MAP_FULL.items():
        m_name = m_full.split(" - ")[1]
        if m_name in refer_str.upper() or refer_str.startswith(f"{m_num} "):
            return m_full
    m = re.search(r'(\d{4})[-/](\d{2})[-/](\d{2})', refer_str)
    if m:
        return MONTH_MAP_FULL.get(int(m.group(2)), "6 - JUNHO")
    m2 = re.search(r'(\d{2})[-/](\d{2})[-/](\d{4})', refer_str)
    if m2:
        return MONTH_MAP_FULL.get(int(m2.group(2)), "6 - JUNHO")
    return refer_str

def verificar_e_atualizar_planilha_njud(drive_service):
    print("\n[Agente] Auditando planilha de jornais (NJUD) no Google Drive...")
    
    wb, temp_path = baixar_planilha_gdrive(drive_service, SPREADSHEET_ID_NJUD, "temp_njud_agente.xlsx")
    if not wb or not temp_path:
        print("[ERRO] Não foi possível baixar a planilha do NJUD via Drive API.")
        return
        
    alterada = False
    
    for s_name in wb.sheetnames:
        if s_name == 'DASHBOARD GERAL':
            continue
        ws = wb[s_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            continue
            
        header = [str(cell).upper().strip() if cell else "" for cell in rows[0]]
        idx_njud = -1
        idx_audio = -1
        idx_refer = -1
        
        for idx, h in enumerate(header):
            if "NJUD" in h or "NOME DO ARQUIVO" in h:
                idx_njud = idx
            elif "AUDIO" in h or "ÁUDIO" in h:
                idx_audio = idx
            elif "REFER" in h or "CAMINHO" in h:
                idx_refer = idx
                
        if idx_njud == -1 or idx_audio == -1:
            continue
            
        for r_idx in range(2, ws.max_row + 1):
            nome_raw = ws.cell(row=r_idx, column=idx_njud + 1).value
            if not nome_raw:
                continue
            nome_arquivo = str(nome_raw).strip()
            if "NJUD" not in nome_arquivo.upper():
                continue
                
            status_atual = ws.cell(row=r_idx, column=idx_audio + 1).value
            if status_atual and any(ok in str(status_atual).upper() for ok in ["OK", "SIM", "✔", "PRONTO"]):
                continue
                
            refer_val = ws.cell(row=r_idx, column=idx_refer + 1).value if idx_refer != -1 else None
            sufixo = obter_sufixo_data_njud(refer_val)
            caminho_col = obter_caminho_mes_njud(refer_val)
            
            nome_final = f"{nome_arquivo} {sufixo}" if sufixo else nome_arquivo
            
            # Caminho físico do arquivo final no Drive 5S
            drive_audio_path_5s = os.path.join(
                NJUD_DRIVE_BASE, "02_AUDIOS_MAILING", 
                f"{obter_caminho_mes_njud_5s(caminho_col)}", f"{nome_final}.mp3"
            ).replace("\\", "/")
            
            # Caminho físico no drive tradicional
            drive_audio_path_trad = os.path.join(
                r"H:\Meu Drive\RADIO TJRN CONTEÚDO\NOT JUDICIARIO (5 MIN)\NJUD 2026",
                caminho_col, "EDITADOS", f"{nome_final}.mp3"
            ).replace("\\", "/")
            
            if os.path.exists(drive_audio_path_5s) or os.path.exists(drive_audio_path_trad):
                ws.cell(row=r_idx, column=idx_audio + 1, value="✔")
                # Console e log seguro
                registrar_log_5s(f"Auditoria NJUD: marcado status [OK] na planilha para {nome_final} (Edição Gerada).")
                alterada = True
                
    if alterada:
        salvar_e_enviar_planilha_gdrive(drive_service, SPREADSHEET_ID_NJUD, wb, temp_path)
        print("[OK] Planilha do NJUD atualizada no Google Drive.")
    else:
        wb.close()
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass
        print("  Nenhum novo áudio pendente de marcação na planilha NJUD.")

def obter_caminho_mes_njud_5s(caminho_col):
    mes_num = 6
    m_mes = re.search(r'(\d+)', caminho_col)
    if m_mes:
        mes_num = int(m_mes.group(1))
    MONTH_MAP_SHORT = {
        1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
        7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ"
    }
    short_name = MONTH_MAP_SHORT.get(mes_num, "JUN")
    return f"{mes_num:02d} - {short_name} - 26"

def run_agent_once(drive_service):
    print("\n" + "="*60)
    print("=== AGENTE DE IA DA RÁDIO TJRN - INICIANDO EXECUÇÃO ===")
    print("="*60)
    
    # 1. Verificar GDrive
    if not verificar_e_montar_drive():
        print("[CRÍTICO] Abortando execução devido a falta do Drive H:.")
        return
        
    # 2. Corrigir inconsistências
    if drive_service:
        analisar_e_corrigir_planilha_boletins(drive_service)
        
    # 3. Executar scripts originais (Fallback)
    executar_pipelines()
    
    # 4. Auditar e fechar planilhas locais
    if drive_service:
        verificar_e_atualizar_planilha_njud(drive_service)
        
    # 5. Enviar relatório consolidado do dia
    email_recipient = carregar_env_var("EMAIL_RECIPIENT", "thi.macedo@gmail.com")
    try:
        from core.send_report import send_daily_report_email
        send_daily_report_email(email_recipient)
    except Exception as e:
        print(f"[AVISO] Falha ao disparar envio do relatório consolidado: {e}")
    
    print("\n" + "="*60)
    print("=== AGENTE DE IA DA RÁDIO TJRN - FIM DA EXECUÇÃO ===")
    print("="*60)

def main():
    if not adquirir_lock():
        sys.exit(1)
        
    parser = argparse.ArgumentParser(description="Agente de IA supervisor da Rádio TJRN.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--once", action="store_true", help="Executa o agente apenas uma vez.")
    group.add_argument("--daemon", action="store_true", help="Executa o agente em loop contínuo.")
    parser.add_argument("--interval", type=int, default=300, help="Intervalo de tempo entre execuções no modo daemon (segundos).")
    parser.add_argument("--watch", action="store_true", help="Ativa o DriveWatcher em background para detecção reativa de novos roteiros.")
    args = parser.parse_args()
    
    # Inicializar serviço Google Drive
    drive_service = obter_google_drive_service()
    if not drive_service:
        print("[CRÍTICO] Falha ao autenticar o serviço da API do Google Drive.")
        sys.exit(1)
        
    if args.once:
        run_agent_once(drive_service)
    elif args.daemon:
        print(f"[Agente] Modo Daemon ativo. Intervalo de execução: {args.interval} segundos.")

        # ---------------------------------------------------------------
        # Watcher reativo (opcional): detecta novos arquivos no Drive
        # ---------------------------------------------------------------
        watcher_thread = None
        if args.watch and drive_watcher_available and drive_service:
            def _on_new_njud(file_meta):
                print(f"[Watcher] Novo roteiro NJUD detectado: {file_meta.get('name')} — disparando pipeline NJUD.")
                _run_pipeline("NJUD", "modules/jornal/njud_pipeline.py")

            def _on_new_giro(file_meta):
                print(f"[Watcher] Novo roteiro Giro detectado: {file_meta.get('name')} — disparando pipeline Giro.")
                _run_pipeline("Giro", "modules/giro/giro_pipeline.py")

            njud_folder_id = carregar_env_var("NJUD_ROTEIROS_FOLDER_ID", "")
            giro_folder_id = carregar_env_var("GIRO_ROTEIROS_FOLDER_ID", "")

            watched = {}
            if njud_folder_id:
                watched[njud_folder_id] = _on_new_njud
            if giro_folder_id:
                watched[giro_folder_id] = _on_new_giro

            if watched:
                watcher = DriveWatcher(
                    service=drive_service,
                    watched_folders=watched,
                    poll_s=120,  # verifica a cada 2 minutos
                )
                watcher_thread = watcher.run_background()
                print(f"[Agente] DriveWatcher ativo — monitorando {len(watched)} pasta(s).")
            else:
                print("[Agente][AVISO] --watch ativado mas nenhum FOLDER_ID configurado no .env. Watcher inativo.")
        elif args.watch and not drive_watcher_available:
            print("[Agente][AVISO] --watch ignorado: módulo core.drive_watcher não disponível.")

        try:
            while True:
                run_agent_once(drive_service)
                print(f"[Agente] Aguardando {args.interval}s para a próxima execução...")
                time.sleep(args.interval)
        except KeyboardInterrupt:
            print("\n[Agente] Parando execução em segundo plano...")
            if watcher_thread and 'watcher' in dir():
                watcher.stop()

if __name__ == "__main__":
    main()
