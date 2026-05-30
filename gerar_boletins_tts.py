import os
import sys
import re
import time
import io
import asyncio
import urllib.request
import pandas as pd
import openpyxl
from datetime import datetime
from pydub import AudioSegment

# Certificar que o caminho do workspace está no python path para importar o processador
workspace_dir = os.path.dirname(os.path.abspath(__file__)).replace("\\", "/")
sys.path.append(workspace_dir)

try:
    from processar_roteiro_completo import limpar_texto_locutor
except ImportError:
    print("[AVISO] Não foi possível importar 'limpar_texto_locutor' de 'processar_roteiro_completo.py'. Usando fallback simples.")
    def limpar_texto_locutor(texto):
        return texto

# Configuração de caminhos e constantes
SPREADSHEET_ID = "1b1xnzvA00H1JC9uTvd6c-PBwQjEzGRs6t_raXG_ztsU"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
DRIVE_DIR = r"H:\Meu Drive\RADIO TJRN CONTEÚDO\0-BOLETINS"
LOCAL_BOLETINS_DIR = os.path.join(workspace_dir, "boletins").replace("\\", "/")

# Mapeamento dos meses no formato NJUD
MONTH_MAP = {
    1: "1 - JANEIRO",
    2: "2 - FEVEREIRO",
    3: "3 - MARÇO",
    4: "4 - ABRIL",
    5: "5 - MAIO",
    6: "6 - JUNHO",
    7: "7 - JULHO",
    8: "8 - AGOSTO",
    9: "9 - SETEMBRO",
    10: "10 - OUTUBRO",
    11: "11 - NOVEMBRO",
    12: "12 - DEZEMBRO"
}

# Carregar assets de áudio
def carregar_audio_asset(caminho, label):
    if os.path.exists(caminho):
        try:
            seg = AudioSegment.from_mp3(caminho)
            print(f"  [ASSET] {label} carregado ({len(seg)}ms)")
            return seg
        except Exception as e:
            print(f"  [ERRO] Falha ao carregar asset {label} ({caminho}): {e}")
    else:
        print(f"  [AVISO] Asset não encontrado: {caminho}")
    return None

# Mapear iniciais do locutor para voz Edge TTS
def obter_config_voz(locutor):
    loc_str = str(locutor).upper() if locutor else ""
    if "LEO" in loc_str:
        return "pt-BR-AntonioNeural", "LEO"
    elif "LIV" in loc_str:
        return "pt-BR-FranciscaNeural", "LIV"
    elif "LET" in loc_str:
        return "pt-BR-FranciscaNeural", "LET"
    elif "SIL" in loc_str:
        return "pt-BR-FranciscaNeural", "SIL"
    else:
        return "pt-BR-FranciscaNeural", "LIV" # Default female voice

# Extrair data de criação a partir das colunas do registro
def extrair_data_registro(nome_edicao, data_criacao, caminho):
    if nome_edicao:
        m = re.search(r'BOLETIM_RADIO_TJRN_(\d{2})_(\d{2})_(\d{4})', nome_edicao, re.IGNORECASE)
        if m:
            return int(m.group(1)), int(m.group(2)), int(m.group(3))
            
    if data_criacao:
        if isinstance(data_criacao, datetime):
            return data_criacao.day, data_criacao.month, data_criacao.year
        if isinstance(data_criacao, str):
            m = re.search(r'(\d{4})-(\d{2})-(\d{2})', data_criacao)
            if m:
                return int(m.group(3)), int(m.group(2)), int(m.group(1))
            m = re.search(r'(\d{2})[/-](\d{2})[/-](\d{4})', data_criacao)
            if m:
                return int(m.group(1)), int(m.group(2)), int(m.group(3))
                
    if caminho:
        m = re.search(r'(\d{2})/(\d{2})', str(caminho))
        if m:
            return int(m.group(1)), int(m.group(2)), 2026
            
    return None, None, None

# Baixar e parsear Google Doc do roteiro
def baixar_e_parsear_roteiro(url):
    m = re.search(r'/d/([a-zA-Z0-9_-]+)', url)
    if not m:
        raise ValueError("URL do Google Doc inválida.")
    doc_id = m.group(1)
    export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=txt"
    
    req = urllib.request.Request(export_url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req) as response:
        content = response.read().decode('utf-8-sig', errors='ignore')
        
    lines = content.split('\n')
    cabeca_lines = []
    off_lines = []
    
    in_cabeca = False
    in_off = False
    
    for line in lines:
        line_strip = line.strip()
        if not line_strip:
            continue
            
        if re.match(r'^(?:cabeça|cabeca|cab)\b', line_strip, re.IGNORECASE):
            in_cabeca = True
            in_off = False
            parts = re.split(r'^(?:cabeça|cabeca|cab)\s*:?\s*', line_strip, flags=re.IGNORECASE)
            if len(parts) > 1 and parts[1].strip():
                cabeca_lines.append(parts[1].strip())
            continue
            
        if re.match(r'^off\b', line_strip, re.IGNORECASE):
            in_off = True
            in_cabeca = False
            parts = re.split(r'^off\s*:?\s*', line_strip, flags=re.IGNORECASE)
            if len(parts) > 1 and parts[1].strip():
                off_lines.append(parts[1].strip())
            continue
            
        if in_cabeca:
            cabeca_lines.append(line_strip)
        elif in_off:
            off_lines.append(line_strip)
            
    cabeca_text = " ".join(cabeca_lines).strip()
    off_text = " ".join(off_lines).strip()
    
    # Fallback simples caso não encontre marcadores estruturados
    if not cabeca_text and not off_text:
        # Tenta dividir ao meio ou assume tudo como OFF
        print("    [AVISO] Roteiro sem marcadores estruturados. Tentando fallback.")
        off_text = " ".join([l.strip() for l in lines if l.strip()]).strip()
        
    return cabeca_text, off_text

# Gerar bytes de áudio via Edge TTS com retry
async def gerar_tts_com_retry(text, voice, rate="+0%"):
    import edge_tts
    for tentativa in range(3):
        try:
            communicate = edge_tts.Communicate(text, voice, rate=rate)
            audio_data = b""
            async for chunk in communicate.stream():
                if chunk["type"] == "audio":
                    audio_data += chunk["data"]
            if audio_data:
                return audio_data
        except Exception as e:
            print(f"      [AVISO] Falha na geração do áudio TTS (tentativa {tentativa+1}): {e}")
            await asyncio.sleep(2)
    raise Exception("Falha de rede ou de serviço com Edge TTS.")

# Mixar Mailing com Trilha de Fundo a 20% do volume original (-14 dB)
def mixar_mailing_com_bg(mailing_audio, bg_audio):
    bg_low = bg_audio - 14  # Redução para 20% do volume
    dur_mailing = len(mailing_audio)
    dur_bg = len(bg_low)
    
    if dur_bg < dur_mailing:
        repeats = (dur_mailing // dur_bg) + 1
        bg_low = bg_low * repeats
        
    bg_low = bg_low[:dur_mailing]
    bg_low = bg_low.fade_in(500).fade_out(800)
    
    return bg_low.overlay(mailing_audio)

# Função central para processar uma pendência
async def processar_boletim(row_idx, row_data, assets, test_mode):
    caminho_col = row_data[0]
    nome_arquivo = row_data[1]
    locutor_col = row_data[3]
    nome_edicao = row_data[6]
    url_doc = row_data[7]
    data_criacao_col = row_data[8]
    
    print(f"\n* Processando linha {row_idx}: {nome_arquivo}")
    
    # 1. Resolver datas e subpastas
    day, month_num, year = extrair_data_registro(nome_edicao, data_criacao_col, caminho_col)
    if not day or not month_num:
        print(f"  [ERRO] Não foi possível resolver a data para a linha {row_idx}.")
        return False
        
    month_name = MONTH_MAP.get(month_num)
    if not month_name:
        print(f"  [ERRO] Mês inválido: {month_num}")
        return False
        
    day_str = str(day).zfill(2)
    
    # 2. Definir caminhos locais de saída
    dia_dir = os.path.join(LOCAL_BOLETINS_DIR, month_name, day_str)
    mailing_dir = os.path.join(dia_dir, "mailing")
    edit_dir = os.path.join(dia_dir, "edit")
    
    os.makedirs(mailing_dir, exist_ok=True)
    os.makedirs(edit_dir, exist_ok=True)
    
    filename_base = nome_edicao if nome_edicao else f"BOLETIM_{day_str}_{month_num}_{year}_{row_idx}"
    txt_saida_path = os.path.join(dia_dir, f"{filename_base}.txt")
    mailing_saida_path = os.path.join(mailing_dir, f"{filename_base}.mp3")
    edit_saida_path = os.path.join(edit_dir, f"{filename_base}.mp3")
    
    # Verificar se já processamos localmente para evitar reprocessamento
    if os.path.exists(mailing_saida_path) and os.path.exists(edit_saida_path) and os.path.exists(txt_saida_path):
        print(f"  - {filename_base} (ignorado, áudios e texto já existem)")
        return True
    
    # 3. Baixar roteiro e parsear
    print(f"  -> Baixando roteiro: {url_doc[:50]}...")
    try:
        cabeca_raw, off_raw = baixar_e_parsear_roteiro(url_doc)
    except Exception as e:
        print(f"  [ERRO] Falha ao baixar ou estruturar roteiro: {e}")
        return False
        
    # Salvar script bruto em txt no diretório do dia
    with open(txt_saida_path, "w", encoding="utf-8") as f:
        f.write(f"CABEÇA:\n{cabeca_raw}\n\nOFF:\n{off_raw}\n")
    print(f"  -> Roteiro em texto salvo em: {txt_saida_path}")
    
    # 4. Limpeza e normalização do texto
    cabeca_limpa = limpar_texto_locutor(cabeca_raw)
    off_limpa = limpar_texto_locutor(off_raw)
    
    # 5. Voz e geração TTS
    voice_name, speaker_name = obter_config_voz(locutor_col)
    print(f"  -> Gravando com a voz '{voice_name}' (Iniciais: {speaker_name})")
    
    try:
        # Gravar Cabeça (rate="+0%" para entonação firme)
        cabeca_bytes = b""
        if cabeca_limpa:
            print("     [TTS] Sintetizando Cabeça...")
            cabeca_bytes = await gerar_tts_com_retry(cabeca_limpa, voice_name, rate="+0%")
            
        # Gravar OFF (rate="+4%" para entonação jornalística)
        off_bytes = b""
        if off_limpa:
            print("     [TTS] Sintetizando OFF...")
            off_bytes = await gerar_tts_com_retry(off_limpa, voice_name, rate="+4%")
    except Exception as e:
        print(f"  [ERRO] Falha na síntese de voz (TTS): {e}")
        return False
        
    # 6. Mixagem de Áudio com Pydub
    try:
        # Converter bytes para AudioSegments
        cabeca_seg = AudioSegment.from_mp3(io.BytesIO(cabeca_bytes)) if cabeca_bytes else AudioSegment.empty()
        off_seg = AudioSegment.from_mp3(io.BytesIO(off_bytes)) if off_bytes else AudioSegment.empty()
        
        # Montar Mailing: Cabeça + Passage Vignette + OFF
        print("  -> Mixando versão Mailing...")
        vht_passagem = assets["vht_passagem"]
        
        # Se houver apenas cabeça ou apenas off, lida elegantemente
        mailing_audio = AudioSegment.empty()
        if cabeca_seg and off_seg:
            mailing_audio = cabeca_seg + vht_passagem + off_seg
        elif cabeca_seg:
            mailing_audio = cabeca_seg
        else:
            mailing_audio = off_seg
            
        # Exportar Mailing a 192k
        mailing_audio.export(mailing_saida_path, format="mp3", bitrate="192k")
        print(f"  [OK] Mailing gerado em: {mailing_saida_path}")
        
        # Montar Editada: Abertura Vignette + (Mailing mixado com BG a 20%) + Encerramento Vignette
        print("  -> Mixando versão Editada (com Trilha de Fundo a 20% e Vinhetas)...")
        vht_abertura = assets["vht_abertura"]
        vht_encerramento = assets["vht_encerramento"]
        bg_boletim = assets["bg_boletim"]
        
        # Mixagem com trilha de fundo
        mixed_speech_bg = mixar_mailing_com_bg(mailing_audio, bg_boletim)
        
        # Concatenação final
        edit_audio = vht_abertura + mixed_speech_bg + vht_encerramento
        
        # Exportar Editada a 192k
        edit_audio.export(edit_saida_path, format="mp3", bitrate="192k")
        print(f"  [OK] Editada gerada em: {edit_saida_path}")
        
        return True
    except Exception as e:
        print(f"  [ERRO] Falha no processamento ou exportação de áudio: {e}")
        return False

# Execução do pipeline principal
async def main():
    import argparse
    parser = argparse.ArgumentParser(description="Centralizador de Gravação Automática de Boletins via TTS.")
    parser.add_argument("--test", action="store_true", help="Executa apenas uma gravação de teste e não altera a planilha final.")
    args = parser.parse_args()
    
    print("=== Processador Central de Boletins Rádio TJRN — Início ===")
    
    # 1. Baixar a planilha controle atualizada
    print(f"\nBaixando planilha de produção a partir do Drive...")
    local_xlsx = "BOLETINS_2026.xlsx"
    try:
        urllib.request.urlretrieve(SHEET_URL, local_xlsx)
        print(f"[OK] Planilha baixada e salva localmente como: {local_xlsx}")
    except Exception as e:
        print(f"[ERRO] Falha ao baixar planilha de controle: {e}")
        sys.exit(1)
        
    # 2. Carregar a planilha com openpyxl para poder fazer atualizações
    wb = openpyxl.load_workbook(local_xlsx)
    
    # 3. Carregar vinhetas e BG do boletim
    vht_abertura_path = os.path.join(LOCAL_BOLETINS_DIR, "VHT/vht_abertura.mp3").replace("\\", "/")
    vht_encerramento_path = os.path.join(LOCAL_BOLETINS_DIR, "VHT/vht_encerramento.mp3").replace("\\", "/")
    vht_passagem_path = os.path.join(LOCAL_BOLETINS_DIR, "VHT/vht_passagem.mp3").replace("\\", "/")
    bg_boletim_path = os.path.join(LOCAL_BOLETINS_DIR, "VHT/bg_boletim.mp3").replace("\\", "/")
    
    print("\nCarregando assets de áudio para boletins...")
    assets = {
        "vht_abertura": carregar_audio_asset(vht_abertura_path, "Abertura"),
        "vht_encerramento": carregar_audio_asset(vht_encerramento_path, "Encerramento"),
        "vht_passagem": carregar_audio_asset(vht_passagem_path, "Passagem"),
        "bg_boletim": carregar_audio_asset(bg_boletim_path, "BG Trilha")
    }
    
    # Verificar se todos os assets vitais estão carregados
    if not all(assets.values()):
        print("[ERRO CRÍTICO] Algum asset de áudio essencial não pôde ser carregado. Abortando.")
        sys.exit(1)
        
    # 4. Listar todas as pendências em todos os meses
    sheets_to_process = [name for name in wb.sheetnames if name != 'DASHBOARD GERAL']
    
    pendencias = [] # Lista de tuplas: (sheet_name, row_idx, row_data)
    
    for s_name in sheets_to_process:
        ws = wb[s_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 5:
            continue
            
        for idx, r in enumerate(rows[5:], start=5):
            if not r[1]: # NOME DO ARQUIVO vazio
                continue
            locutor = r[3]
            editor = r[4]
            url = r[7]
            
            # Validação se a URL existe para download
            if not url or 'document/d/' not in str(url):
                continue
                
            is_loc_pending = not locutor or '✔' not in str(locutor)
            is_edit_pending = not editor or '✔' not in str(editor)
            
            if is_loc_pending or is_edit_pending:
                pendencias.append((s_name, idx, r))
                
    if not pendencias:
        print("\n[INFO] Nenhuma pendência encontrada na planilha! Tudo finalizado.")
        sys.exit(0)
        
    print(f"\nTotal de pendências detectadas: {len(pendencias)}")
    
    if args.test:
        print("\n*** MODO DE TESTE ATIVO — Processando apenas 1 pendência ***")
        pendencias = pendencias[:1]
        
    sucessos = 0
    linhas_atualizadas = [] # Armazena quais linhas foram modificadas
    
    # Limitar concorrência para não estourar limite do Edge TTS
    sem = asyncio.Semaphore(4)
    
    async def processar_com_sem(sheet_name, row_idx, row_data):
        async with sem:
            try:
                processado = await processar_boletim(row_idx, row_data, assets, args.test)
                if processado:
                    # Pequeno delay pós-sucesso dentro do worker para espaçar requisições
                    await asyncio.sleep(1.0)
                    return sheet_name, row_idx, row_data
            except Exception as e:
                print(f"  [ERRO] Falha crítica ao processar linha {row_idx} ({sheet_name}): {e}")
            return None

    print(f"Iniciando processamento concorrente de {len(pendencias)} pendências...")
    tasks = [processar_com_sem(s_name, r_idx, r_data) for s_name, r_idx, r_data in pendencias]
    results = await asyncio.gather(*tasks)
    
    for res in results:
        if res:
            sucessos += 1
            if not args.test:
                sheet_name, row_idx, row_data = res
                loc_val = row_data[3]
                _, speaker_init = obter_config_voz(loc_val)
                novo_locutor_texto = f"{speaker_init} ✔"
                novo_editor_texto = "LET ✔"
                linhas_atualizadas.append((sheet_name, row_idx, novo_locutor_texto, novo_editor_texto))
            
    print(f"\n=== PROCESSAMENTO FINALIZADO: {sucessos} de {len(pendencias)} concluídos ===")
    
    # 5. Se foi gravado com sucesso e não é teste, atualizar planilhas
    if not args.test and sucessos > 0:
        print("\nAtualizando planilhas com as novas confirmações de gravação...")
        for s_name, r_idx, loc_txt, edit_txt in linhas_atualizadas:
            ws = wb[s_name]
            # No openpyxl, a linha é 1-indexed. Como iter_rows retornou a linha 0-indexed,
            # o row_idx que passamos (inicializado em 5 do start=5) já bate exatamente com o Excel físico!
            ws.cell(row=r_idx + 1, column=4, value=loc_txt) # Coluna D: LOCUTOR
            ws.cell(row=r_idx + 1, column=5, value=edit_txt) # Coluna E: EDITOR
            
        # Salvar planilha local de boletins
        local_updated_xlsx = os.path.join(LOCAL_BOLETINS_DIR, "BOLETINS_2026_ATUALIZADO.xlsx").replace("\\", "/")
        wb.save(local_updated_xlsx)
        print(f"Planilha Excel atualizada salva localmente em: {local_updated_xlsx}")
        
        # Exportar cada aba para CSV
        csv_dir = os.path.join(LOCAL_BOLETINS_DIR, "planilha_csv").replace("\\", "/")
        os.makedirs(csv_dir, exist_ok=True)
        
        print("\nExportando abas atualizadas para arquivos CSV...")
        for s_name in sheets_to_process:
            ws = wb[s_name]
            rows_data = []
            for r in ws.iter_rows(values_only=True):
                # Mantém apenas as linhas onde há alguma célula preenchida
                if any(r):
                    rows_data.append(r)
            # Criar DataFrame e salvar como CSV
            df = pd.DataFrame(rows_data[4:]) # Pula a firula inicial do Dashboard da aba e usa a linha 4 como dados
            df.columns = rows_data[4] # Seta headers corretos
            # Remover a primeira linha de dados se ela duplicar os headers
            df = df.iloc[1:]
            
            csv_path = os.path.join(csv_dir, f"{s_name}.csv").replace("\\", "/")
            df.to_csv(csv_path, index=False, encoding="utf-8-sig")
            print(f"  [OK] Exportado CSV: {csv_path}")
            
        # 6. Copiar de volta para a pasta do Google Drive (H:\Meu Drive\RADIO TJRN CONTEÚDO\0-BOLETINS)
        print(f"\nSincronizando planilhas atualizadas com a pasta do Google Drive em '{DRIVE_DIR}'...")
        if os.path.exists(DRIVE_DIR):
            drive_xlsx_path = os.path.join(DRIVE_DIR, "BOLETINS_2026_ATUALIZADO.xlsx").replace("\\", "/")
            drive_csv_dir = os.path.join(DRIVE_DIR, "planilha_csv").replace("\\", "/")
            os.makedirs(drive_csv_dir, exist_ok=True)
            
            # Copiar Excel
            import shutil
            shutil.copy2(local_updated_xlsx, drive_xlsx_path)
            print(f"  [OK] Planilha Excel copiada para o Drive: {drive_xlsx_path}")
            
            # Copiar todos os CSVs
            for file in os.listdir(csv_dir):
                if file.endswith(".csv"):
                    shutil.copy2(os.path.join(csv_dir, file), os.path.join(drive_csv_dir, file))
            print(f"  [OK] Todos os CSVs copiados para o Drive em: {drive_csv_dir}")
        else:
            print(f"[AVISO] Pasta do Drive '{DRIVE_DIR}' não encontrada. Planilhas salvas apenas localmente.")
            
        # Sincronizar os áudios e textos gerados com as pastas do Drive
        try:
            from sincronizar_boletins_drive import sincronizar
            sincronizar()
        except Exception as e:
            print(f"[ERRO] Falha ao executar sincronização com o Drive: {e}")
            
    print("\n=== PIPELINE CONCLUÍDO ===")

if __name__ == "__main__":
    asyncio.run(main())
