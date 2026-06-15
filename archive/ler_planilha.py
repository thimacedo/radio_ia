import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def main():
    wb = openpyxl.load_workbook("BOLETINS_2026.xlsx", data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name == 'DASHBOARD GERAL':
            continue
        ws = wb[sheet_name]
        locutores = set()
        for row in list(ws.iter_rows(values_only=True))[5:]:
            if len(row) > 3 and row[3]:
                locutores.add(row[3])
        print(f"Sheet: {sheet_name}, Locutores: {locutores}")

if __name__ == "__main__":
    main()
