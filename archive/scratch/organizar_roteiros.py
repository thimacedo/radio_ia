# -*- coding: utf-8 -*-
import os
import sys
import re
import urllib.request
import json
import openpyxl
from pathlib import Path
from datetime import datetime

# Configurar caminhos do projeto
current_dir = Path(__file__).parent.resolve()
project_root = current_dir.parent.parent.resolve()
sys.path.append(str(project_root))

try:
    from core.best_practices import carregar_env_var
    from core.constants import MONTH_MAP_SHORT, MONTH_MAP_FULL, WEEKDAYS_PT
except ImportError:
    def carregar_env_var(chave, fallback):
        return fallback
    MONTH_MAP_SHORT = {
        1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
        7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ"
    }
    MONTH_MAP_FULL = {
        1: "1 - JANEIRO", 2: "2 - FEVEREIRO", 3: "3 - MARÇO", 4: "4 - ABRIL",
        5: "5 - MAIO",    6: "6 - JUNHO",     7: "7 - JULHO",  8: "8 - AGOSTO",
        9: "9 - SETEMBRO",10: "10 - OUTUBRO", 11: "11 - NOVEMBRO", 12: "12 - DEZEMBRO"
    }
    WEEKDAYS_PT = {0: "SEG", 1: "TER", 2: "QUA", 3: "QUI", 4: "SEX", 5: "SAB", 6: "DOM"}

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
except ImportError as exc:
    raise ImportError("google-api-python-client e google-auth-oauthlib são necessários.")

# Configurações do Drive e Planilha
DRIVE_ROOT = Path(carregar_env_var("DRIVE_ROOT", "H:/Meu Drive/RADIO TJRN CONTEÚDO"))
DRIVE_ROTEIROS_BASE = DRIVE_ROOT / "00_PRODUCAO_2026" / "01_BOLETINS_DIARIOS" / "01_ROTEIROS"

SPREADSHEET_ID = "1b1xnzvA00H1JC9uTvd6c-PBwQjEzGRs6t_raXG_ztsU"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
LOCAL_XLSX = project_root / "modules" / "boletins" / "BOLETINS_ATUAL.xlsx"

DEFAULT_CREDS_REL = carregar_env_var("GOOGLE_APPLICATION_CREDENTIALS", "config/credentials/gen-lang-client-0980378916-8cc8eb1488d1.json")
CREDENTIALS_PATH = project_root / DEFAULT_CREDS_REL

MONTH_ABBREV_MAP = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12
}

def obter_dia_semana_pt(ano, mes, dia):
    try:
        dt = datetime(ano, mes, dia)
        return WEEKDAYS_PT[dt.weekday()]
    except Exception:
        return "SEG"

def extrair_data_registro(nome_edicao, data_criacao, caminho):
    if nome_edicao:
        m = re.search(r'BOLETIM_RADIO_TJRN_(\d{2})_(\d{2})_(\d{4})', str(nome_edicao), re.IGNORECASE)
        if m:
            return int(m.group(1)), int(m.group(2)), int(m.group(3))
            
    if data_criacao:
        if isinstance(data_criacao, datetime):
            return data_criacao.day, data_criacao.month, data_criacao.year
        if isinstance(data_criacao, str):
            m = re.search(r'(\d{4})-(\d{2})-(\d{2})', data_criacao)
            if m:
                return int(m.group(3)), int(m.group(2)), int(m.group(1))
            m = re.search(r'(\d{2})[/-](\d{2})[/-](\d{4})', data_criacao)
            if m:
                return int(m.group(1)), int(m.group(2)), int(m.group(3))
                
    if caminho:
        m = re.search(r'(\d{2})/(\d{2})', str(caminho))
        if m:
            return int(m.group(1)), int(m.group(2)), 2026
            
    return None, None, None

def _build_drive_service(credentials_path: Path):
    creds = service_account.Credentials.from_service_account_file(
        str(credentials_path), scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)

def obter_id_por_caminho(service, path_parts):
    current_parent = 'root'
    for part in path_parts:
        query = f"name = '{part}' and mimeType = 'application/vnd.google-apps.folder' and '{current_parent}' in parents and trashed = false"
        if current_parent == 'root':
            res = service.files().list(q=query, fields="files(id)").execute()
            files = res.get("files", [])
            if not files:
                query_shared = f"name = '{part}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
                res = service.files().list(q=query_shared, fields="files(id)").execute()
                files = res.get("files", [])
        else:
            res = service.files().list(q=query, fields="files(id)").execute()
            files = res.get("files", [])
            
        if not files:
            return None
        current_parent = files[0]["id"]
    return current_parent

def obter_ou_criar_pasta_drive(service, parent_id, folder_name):
    query = f"name = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder' and '{parent_id}' in parents and trashed = false"
    results = service.files().list(q=query, fields="files(id)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]
    
    metadata = {
        'name': folder_name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }
    folder = service.files().create(body=metadata, fields='id').execute()
    return folder.get('id')

def main():
    print("=== Sincronizador de Roteiros (Via Google Drive API) ===")
    
    if not CREDENTIALS_PATH.exists():
        print(f"[ERRO] Credenciais do Drive não encontradas em {CREDENTIALS_PATH}")
        sys.exit(1)
        
    service = _build_drive_service(CREDENTIALS_PATH)
    
    # 1. Encontrar o ID da pasta 01_ROTEIROS no Drive
    roteiros_root_id = obter_id_por_caminho(service, ["00_PRODUCAO_2026", "01_BOLETINS_DIARIOS", "01_ROTEIROS"])
    if not roteiros_root_id:
        print("[ERRO] Não foi possível encontrar a pasta '01_ROTEIROS' no Google Drive.")
        sys.exit(1)
    print(f"Pasta raiz '01_ROTEIROS' encontrada com ID: {roteiros_root_id}")
    
    print("Baixando planilha de Boletins...")
    try:
        urllib.request.urlretrieve(SHEET_URL, LOCAL_XLSX)
        print("Planilha de Boletins baixada com sucesso.")
    except Exception as e:
        print(f"[ERRO] Falha ao baixar planilha: {e}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(LOCAL_XLSX, data_only=True)
    
    month_sheets = []
    for s_name in wb.sheetnames:
        s_upper = s_name.upper()
        m = re.match(r'^([A-Z]{3})\d{4}$', s_upper)
        if m and m.group(1) in MONTH_ABBREV_MAP:
            month_sheets.append((s_name, MONTH_ABBREV_MAP[m.group(1)]))
        elif s_upper in [
            "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", 
            "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"
        ]:
            full_months = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]
            month_sheets.append((s_name, full_months.index(s_upper) + 1))
            
    total_verificados = 0
    total_criados = 0
    
    # Cache local de pastas de meses e dias para evitar chamadas redundantes à API
    cache_month_ids = {}
    cache_day_ids = {}
    
    for s_name, mes_padrao in month_sheets:
        print(f"\nProcessando aba '{s_name}' (Mês {mes_padrao})...")
        ws = wb[s_name]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            if not row or not any(row):
                continue
                
            caminho_col = row[0]
            nome_edicao = row[6]
            url_doc = row[7]
            data_criacao = row[8]
            
            if not url_doc or not isinstance(url_doc, str) or "docs.google.com" not in url_doc:
                continue
                
            m_id = re.search(r'/d/([a-zA-Z0-9_-]{25,})', url_doc)
            if not m_id:
                continue
            doc_id = m_id.group(1)
            
            day, month, year = extrair_data_registro(nome_edicao, data_criacao, caminho_col)
            if not day:
                if caminho_col and re.match(r'^\d+$', str(caminho_col).strip()):
                    day = int(str(caminho_col).strip())
                    month = mes_padrao
                    year = 2026
                else:
                    continue
            
            if not month:
                month = mes_padrao
            if not year:
                year = 2026
                
            short_month_name = MONTH_MAP_SHORT.get(month, "JUN")
            month_folder_name = f"{month:02d} - {short_month_name} - 26"
            
            dia_semana = obter_dia_semana_pt(year, month, day)
            dia_folder_name = f"{day:02d} {month:02d} - {dia_semana}"
            
            # 2. Obter ou criar pasta de Mês
            if month_folder_name not in cache_month_ids:
                month_folder_id = obter_ou_criar_pasta_drive(service, roteiros_root_id, month_folder_name)
                cache_month_ids[month_folder_name] = month_folder_id
            else:
                month_folder_id = cache_month_ids[month_folder_name]
                
            # 3. Obter ou criar pasta de Dia
            day_key = (month_folder_name, dia_folder_name)
            if day_key not in cache_day_ids:
                day_folder_id = obter_ou_criar_pasta_drive(service, month_folder_id, dia_folder_name)
                cache_day_ids[day_key] = day_folder_id
            else:
                day_folder_id = cache_day_ids[day_key]
                
            file_base_name = nome_edicao if nome_edicao else f"BOLETIM_RADIO_TJRN_{day:02d}_{month:02d}_{year}_L{row_idx}"
            file_base_name = str(file_base_name).strip()
            
            total_verificados += 1
            
            # 4. Verificar se o atalho do roteiro já existe na pasta do dia no Drive
            query_shortcut = f"name = '{file_base_name}' and '{day_folder_id}' in parents and trashed = false"
            res_sh = service.files().list(q=query_shortcut, fields="files(id)").execute()
            
            if not res_sh.get("files"):
                print(f"  [FALTANDO NO DRIVE] {month_folder_name}/{dia_folder_name}/{file_base_name}")
                try:
                    file_metadata = {
                        'name': file_base_name,
                        'mimeType': 'application/vnd.google-apps.shortcut',
                        'parents': [day_folder_id],
                        'shortcutDetails': {
                            'targetId': doc_id
                        }
                    }
                    service.files().create(body=file_metadata).execute()
                    print(f"    [CRIADO] Atalho de Roteiro inserido na nuvem.")
                    total_criados += 1
                except Exception as e_write:
                    print(f"    [ERRO] Falha ao criar atalho no Drive: {e_write}")
            
    wb.close()
    
    print("\n" + "="*50)
    print(f"Execução Finalizada via Cloud API.")
    print(f"Total de Roteiros Verificados: {total_verificados}")
    print(f"Total de Atalhos de Roteiro Criados na Nuvem: {total_criados}")
    print("="*50)

if __name__ == "__main__":
    main()
